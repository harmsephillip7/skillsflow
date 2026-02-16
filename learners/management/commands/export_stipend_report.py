"""
Management command to generate stipend payment reports and export to CSV/Excel.

Usage:
    python manage.py export_stipend_report --month 12 --year 2025 --format csv
    python manage.py export_stipend_report --month 12 --year 2025 --format excel --status APPROVED
    python manage.py export_stipend_report --output /path/to/report.xlsx
"""
import csv
from django.core.management.base import BaseCommand, CommandError
from django.db.models import Sum, Count, Q
from datetime import date
from decimal import Decimal
from pathlib import Path

from learners.models import StipendCalculation


class Command(BaseCommand):
    help = 'Export stipend payment report to CSV or Excel'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--month',
            type=int,
            help='Month to export (1-12). Defaults to previous month.'
        )
        parser.add_argument(
            '--year',
            type=int,
            help='Year to export. Defaults to current year.'
        )
        parser.add_argument(
            '--format',
            choices=['csv', 'excel'],
            default='csv',
            help='Export format (default: csv)'
        )
        parser.add_argument(
            '--output',
            type=str,
            help='Output file path. If not specified, uses default naming.'
        )
        parser.add_argument(
            '--status',
            choices=['CALCULATED', 'VERIFIED', 'APPROVED', 'PAID'],
            help='Filter by stipend status'
        )
        parser.add_argument(
            '--campus-id',
            type=int,
            help='Filter by campus ID'
        )
        parser.add_argument(
            '--include-unverified',
            action='store_true',
            help='Include stipends with incomplete verification (default: exclude)'
        )
    
    def handle(self, *args, **options):
        # Determine period
        month = options.get('month')
        year = options.get('year')
        
        if not month or not year:
            today = date.today()
            if month is None:
                month = today.month - 1
                if month == 0:
                    month = 12
                    year = today.year - 1 if year is None else year
            if year is None:
                year = today.year
        
        if not 1 <= month <= 12:
            raise CommandError(f'Invalid month: {month}. Must be between 1 and 12.')
        
        period_name = date(year, month, 1).strftime('%B %Y')
        
        self.stdout.write(f'\n{self.style.HTTP_INFO}="=" * 60}')
        self.stdout.write(f'{self.style.HTTP_INFO}Exporting Stipend Report for {period_name}')
        self.stdout.write(f'{self.style.HTTP_INFO}="=" * 60}\n')
        
        # Build queryset
        stipends = StipendCalculation.objects.filter(
            month=month,
            year=year
        ).select_related(
            'placement', 'placement__learner', 'placement__host__employer',
            'approved_by'
        ).order_by('placement__learner__last_name', 'placement__learner__first_name')
        
        # Apply filters
        if options.get('status'):
            stipends = stipends.filter(status=options['status'])
        
        if options.get('campus_id'):
            stipends = stipends.filter(placement__learner__campus_id=options['campus_id'])
        
        if not options.get('include_unverified'):
            # Update verification stats first
            for stipend in stipends:
                stipend.update_verification_stats()
            
            # Filter for 100% verified only
            stipends = [s for s in stipends if s.can_finalize]
        
        count = len(stipends) if isinstance(stipends, list) else stipends.count()
        
        if count == 0:
            self.stdout.write(self.style.WARNING('No stipends found matching criteria.'))
            return
        
        self.stdout.write(f'Found {count} stipend record(s).\n')
        
        # Determine output path
        output_path = options.get('output')
        if not output_path:
            format_ext = 'xlsx' if options['format'] == 'excel' else 'csv'
            safe_period = period_name.replace(' ', '_')
            output_path = f'stipend_report_{safe_period}.{format_ext}'
        
        output_path = Path(output_path)
        
        # Export data
        if options['format'] == 'csv':
            self._export_csv(stipends, output_path, period_name)
        else:
            self._export_excel(stipends, output_path, period_name)
        
        # Summary statistics
        if isinstance(stipends, list):
            total_amount = sum(s.net_amount for s in stipends)
            total_verified = sum(1 for s in stipends if s.can_finalize)
        else:
            aggregates = stipends.aggregate(
                total_amount=Sum('net_amount'),
                avg_amount=Sum('net_amount') / Count('id')
            )
            total_amount = aggregates['total_amount'] or Decimal('0')
            total_verified = sum(1 for s in stipends if s.can_finalize)
        
        self.stdout.write(f'\n{self.style.HTTP_INFO}="=" * 60}')
        self.stdout.write(f'{self.style.HTTP_INFO}Summary')
        self.stdout.write(f'{self.style.HTTP_INFO}="=" * 60}')
        self.stdout.write(f'Total records: {count}')
        self.stdout.write(f'Fully verified: {total_verified}')
        self.stdout.write(f'Total amount: {self.style.SUCCESS(f"R{total_amount:,.2f}")}')
        self.stdout.write(f'\n{self.style.SUCCESS(f"✓ Report exported to: {output_path}")}\n')
    
    def _export_csv(self, stipends, output_path, period_name):
        """Export stipends to CSV file"""
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            # Header
            writer.writerow(['Stipend Payment Report'])
            writer.writerow([f'Period: {period_name}'])
            writer.writerow([f'Generated: {date.today().strftime("%Y-%m-%d")}'])
            writer.writerow([])  # Empty row
            
            # Column headers
            writer.writerow([
                'ID',
                'Learner Name',
                'ID Number',
                'Host Employer',
                'Days Present',
                'Annual Leave',
                'Sick Leave',
                'Family Leave',
                'Unpaid Leave',
                'Public Holiday',
                'Absent',
                'Suspended',
                'Total Paid Days',
                'Daily Rate',
                'Gross Amount',
                'Deductions',
                'Net Amount',
                'Status',
                'Verification %',
                'Can Finalize',
                'Approved By',
                'Approved Date',
            ])
            
            # Data rows
            for stipend in stipends:
                learner = stipend.placement.learner
                writer.writerow([
                    stipend.id,
                    learner.get_full_name(),
                    learner.sa_id_number or '',
                    stipend.placement.host.employer.name,
                    stipend.days_present,
                    stipend.days_annual_leave,
                    stipend.days_sick_leave,
                    stipend.days_family_leave,
                    stipend.days_unpaid_leave,
                    stipend.days_public_holiday,
                    stipend.days_absent,
                    stipend.days_suspended,
                    stipend.paid_days,
                    f'{stipend.daily_rate:.2f}',
                    f'{stipend.gross_amount:.2f}',
                    f'{stipend.total_deductions:.2f}',
                    f'{stipend.net_amount:.2f}',
                    stipend.status,
                    f'{stipend.verification_percentage:.1f}%',
                    'Yes' if stipend.can_finalize else 'No',
                    stipend.approved_by.get_full_name() if stipend.approved_by else '',
                    stipend.approved_at.strftime('%Y-%m-%d %H:%M') if stipend.approved_at else '',
                ])
        
        self.stdout.write(f'CSV file created: {output_path}')
    
    def _export_excel(self, stipends, output_path, period_name):
        """Export stipends to Excel file"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise CommandError(
                'openpyxl is required for Excel export. Install with: pip install openpyxl'
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Stipend Report'
        
        # Title
        ws['A1'] = 'Stipend Payment Report'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f'Period: {period_name}'
        ws['A3'] = f'Generated: {date.today().strftime("%Y-%m-%d")}'
        
        # Column headers
        headers = [
            'ID', 'Learner Name', 'ID Number', 'Host Employer',
            'Days Present', 'Annual Leave', 'Sick Leave', 'Family Leave',
            'Unpaid Leave', 'Public Holiday', 'Absent', 'Suspended',
            'Total Paid Days', 'Daily Rate', 'Gross Amount', 'Deductions',
            'Net Amount', 'Status', 'Verification %', 'Can Finalize',
            'Approved By', 'Approved Date'
        ]
        
        header_row = 5
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, stipend in enumerate(stipends, start=header_row + 1):
            learner = stipend.placement.learner
            data = [
                stipend.id,
                learner.get_full_name(),
                learner.sa_id_number or '',
                stipend.placement.host.employer.name,
                stipend.days_present,
                stipend.days_annual_leave,
                stipend.days_sick_leave,
                stipend.days_family_leave,
                stipend.days_unpaid_leave,
                stipend.days_public_holiday,
                stipend.days_absent,
                stipend.days_suspended,
                stipend.paid_days,
                float(stipend.daily_rate),
                float(stipend.gross_amount),
                float(stipend.total_deductions),
                float(stipend.net_amount),
                stipend.status,
                f'{stipend.verification_percentage:.1f}%',
                'Yes' if stipend.can_finalize else 'No',
                stipend.approved_by.get_full_name() if stipend.approved_by else '',
                stipend.approved_at.strftime('%Y-%m-%d %H:%M') if stipend.approved_at else '',
            ]
            
            for col, value in enumerate(data, start=1):
                ws.cell(row=row_idx, column=col, value=value)
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        wb.save(output_path)
        self.stdout.write(f'Excel file created: {output_path}')
