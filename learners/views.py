"""
Learner Management Views
Includes Kanban, Pivot, Timetable, Performance, and Import/Export functionality
"""
import csv
import json
import io
from datetime import date, datetime, timedelta
from decimal import Decimal
from collections import defaultdict

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, StreamingHttpResponse
from django.views import View
from django.views.generic import ListView, DetailView, TemplateView
from django.db.models import Count, Avg, Q, F, Sum, Case, When, Value, CharField
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.utils import timezone

from .models import Learner, Address, SETA
from academics.models import Enrollment, Qualification, Module
from assessments.models import AssessmentResult, AssessmentActivity
from tenants.models import Campus, Brand
from core.context_processors import get_selected_campus


# =============================================================================
# KANBAN VIEW - Drag & Drop Learner Management
# =============================================================================

class LearnerKanbanView(LoginRequiredMixin, TemplateView):
    """
    Kanban board showing learners organized by enrollment status
    Supports drag-and-drop status changes
    """
    template_name = 'learners/kanban.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters
        qualification_id = self.request.GET.get('qualification')
        campus_id = self.request.GET.get('campus')
        search = self.request.GET.get('search', '')
        
        # Base queryset
        enrollments = Enrollment.objects.select_related(
            'learner', 'qualification', 'campus'
        ).order_by('-enrollment_date')
        
        # Apply filters
        if qualification_id:
            enrollments = enrollments.filter(qualification_id=qualification_id)
        if campus_id:
            enrollments = enrollments.filter(campus_id=campus_id)
        else:
            # Apply global campus filter if set
            selected_campus = get_selected_campus(self.request)
            if selected_campus:
                enrollments = enrollments.filter(campus=selected_campus)
        if search:
            enrollments = enrollments.filter(
                Q(learner__first_name__icontains=search) |
                Q(learner__last_name__icontains=search) |
                Q(learner__sa_id_number__icontains=search) |
                Q(enrollment_number__icontains=search)
            )
        
        # Define Kanban columns with colors
        kanban_columns = [
            {'status': 'APPLIED', 'label': 'Applied', 'color': 'slate', 'icon': 'inbox'},
            {'status': 'DOC_CHECK', 'label': 'Document Check', 'color': 'amber', 'icon': 'document-magnifying-glass'},
            {'status': 'REGISTERED', 'label': 'Registered', 'color': 'blue', 'icon': 'clipboard-document-check'},
            {'status': 'ENROLLED', 'label': 'Enrolled', 'color': 'indigo', 'icon': 'academic-cap'},
            {'status': 'ACTIVE', 'label': 'Active', 'color': 'green', 'icon': 'play-circle'},
            {'status': 'ON_HOLD', 'label': 'On Hold', 'color': 'orange', 'icon': 'pause-circle'},
            {'status': 'COMPLETED', 'label': 'Completed', 'color': 'teal', 'icon': 'check-circle'},
            {'status': 'CERTIFIED', 'label': 'Certified', 'color': 'emerald', 'icon': 'trophy'},
        ]
        
        # Group enrollments by status
        for column in kanban_columns:
            column['enrollments'] = enrollments.filter(status=column['status'])[:50]
            column['count'] = enrollments.filter(status=column['status']).count()
        
        # Get counts for hidden statuses
        withdrawn_count = enrollments.filter(status='WITHDRAWN').count()
        transferred_count = enrollments.filter(status='TRANSFERRED').count()
        expired_count = enrollments.filter(status='EXPIRED').count()
        
        context.update({
            'kanban_columns': kanban_columns,
            'withdrawn_count': withdrawn_count,
            'transferred_count': transferred_count,
            'expired_count': expired_count,
            'qualifications': Qualification.objects.filter(is_active=True),
            'campuses': Campus.objects.filter(is_active=True),
            'selected_qualification': qualification_id,
            'selected_campus': campus_id,
            'search': search,
            'total_enrollments': enrollments.count(),
        })
        
        return context


@login_required
def kanban_update_status(request):
    """AJAX endpoint to update enrollment status from Kanban drag-drop"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            enrollment_id = data.get('enrollment_id')
            new_status = data.get('new_status')
            
            enrollment = get_object_or_404(Enrollment, id=enrollment_id)
            old_status = enrollment.status
            
            # Update status
            enrollment.status = new_status
            enrollment.status_changed_at = timezone.now()
            enrollment.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Status updated from {old_status} to {new_status}',
                'enrollment_id': enrollment_id,
                'new_status': new_status
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


# =============================================================================
# PIVOT TABLE VIEWS - Analytics & Reporting
# =============================================================================

class LearnerPivotView(LoginRequiredMixin, TemplateView):
    """
    Pivot table view for learner analytics
    Group by: qualification, campus, status, funding type, gender, etc.
    """
    template_name = 'learners/pivot.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get pivot configuration
        row_field = self.request.GET.get('rows', 'qualification')
        col_field = self.request.GET.get('cols', 'status')
        value_field = self.request.GET.get('value', 'count')
        
        # Define available fields
        pivot_fields = {
            'qualification': {'label': 'Qualification', 'model_field': 'qualification__short_title'},
            'campus': {'label': 'Campus', 'model_field': 'campus__name'},
            'status': {'label': 'Status', 'model_field': 'status'},
            'funding_type': {'label': 'Funding Type', 'model_field': 'funding_type'},
            'gender': {'label': 'Gender', 'model_field': 'learner__gender'},
            'population_group': {'label': 'Population Group', 'model_field': 'learner__population_group'},
            'province': {'label': 'Province', 'model_field': 'learner__province_code'},
            'year': {'label': 'Enrollment Year', 'model_field': 'enrollment_date__year'},
        }
        
        # Build pivot data
        row_model_field = pivot_fields.get(row_field, {}).get('model_field', 'qualification__short_title')
        col_model_field = pivot_fields.get(col_field, {}).get('model_field', 'status')
        
        # Get unique column values (column headers)
        column_headers = list(Enrollment.objects.values_list(col_model_field, flat=True).distinct().order_by(col_model_field))
        # Replace None with readable value
        column_headers = [c if c else '(None)' for c in column_headers]
        
        # Build pivot query
        pivot_query = Enrollment.objects.values(
            row_model_field, col_model_field
        ).annotate(
            count=Count('id')
        ).order_by(row_model_field)
        
        # Restructure for template - dictionary format for .items() iteration
        pivot_data = {}  # {row_key: {col1: count, col2: count, 'total': total}}
        column_totals = defaultdict(int)
        grand_total = 0
        
        for item in pivot_query:
            row_key = item[row_model_field] or '(None)'
            col_key = item[col_model_field] or '(None)'
            count = item['count']
            
            if row_key not in pivot_data:
                pivot_data[row_key] = {'total': 0}
            
            pivot_data[row_key][col_key] = count
            pivot_data[row_key]['total'] += count
            column_totals[col_key] += count
            grand_total += count
        
        # Get active count for summary card
        active_count = Enrollment.objects.filter(status='ACTIVE').count()
        completed_count = Enrollment.objects.filter(status__in=['COMPLETED', 'CERTIFIED']).count()
        
        context.update({
            'pivot_fields': pivot_fields,
            'rows': row_field,
            'cols': col_field,
            'values': value_field,
            'column_headers': column_headers,
            'pivot_data': pivot_data,
            'column_totals': dict(column_totals),
            'grand_total': grand_total,
            'row_label': pivot_fields.get(row_field, {}).get('label', 'Row'),
            'col_label': pivot_fields.get(col_field, {}).get('label', 'Column'),
            'active_count': active_count,
            'completed_count': completed_count,
        })
        
        return context


# =============================================================================
# TIMETABLE VIEWS
# =============================================================================

class TimetableView(LoginRequiredMixin, TemplateView):
    """
    Timetable view showing class schedules
    """
    template_name = 'learners/timetable.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters
        qualification_id = self.request.GET.get('qualification')
        campus_id = self.request.GET.get('campus')
        week_offset = int(self.request.GET.get('week', 0))
        
        # Calculate week dates
        today = date.today()
        start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
        week_dates = [start_of_week + timedelta(days=i) for i in range(5)]  # Mon-Fri
        
        # Time slots (8 AM to 5 PM)
        time_slots = [
            {'start': '08:00', 'end': '09:00'},
            {'start': '09:00', 'end': '10:00'},
            {'start': '10:00', 'end': '11:00'},
            {'start': '11:00', 'end': '12:00'},
            {'start': '12:00', 'end': '13:00'},
            {'start': '13:00', 'end': '14:00'},
            {'start': '14:00', 'end': '15:00'},
            {'start': '15:00', 'end': '16:00'},
            {'start': '16:00', 'end': '17:00'},
        ]
        
        # Sample timetable data structure
        # In production, this would come from a Timetable model
        timetable_data = self._generate_sample_timetable(week_dates, time_slots, qualification_id)
        
        context.update({
            'week_dates': week_dates,
            'time_slots': time_slots,
            'timetable_data': timetable_data,
            'week_offset': week_offset,
            'start_of_week': start_of_week,
            'qualifications': Qualification.objects.filter(is_active=True),
            'campuses': Campus.objects.filter(is_active=True),
            'selected_qualification': qualification_id,
            'selected_campus': campus_id,
        })
        
        return context
    
    def _generate_sample_timetable(self, week_dates, time_slots, qualification_id=None):
        """Generate sample timetable data"""
        # Get modules for display
        modules = Module.objects.all()[:10]
        if qualification_id:
            modules = Module.objects.filter(qualification_id=qualification_id)[:10]
        
        timetable = {}
        colors = ['blue', 'green', 'purple', 'orange', 'teal', 'pink', 'indigo']
        
        import random
        for i, day in enumerate(week_dates):
            day_key = day.strftime('%Y-%m-%d')
            timetable[day_key] = {}
            
            # Add 2-3 classes per day
            random.seed(day.day + i)  # Consistent per day
            
            used_slots = set()
            for j, module in enumerate(list(modules)[:3]):
                # Pick a random time slot
                available_slots = [s for idx, s in enumerate(time_slots) if idx not in used_slots and idx < 7]
                if available_slots:
                    slot = random.choice(available_slots)
                    slot_idx = time_slots.index(slot)
                    used_slots.add(slot_idx)
                    used_slots.add(slot_idx + 1)  # Block next slot too
                    
                    timetable[day_key][slot['start']] = {
                        'module': module.title[:30] + '...' if len(module.title) > 30 else module.title,
                        'module_code': module.code,
                        'venue': f'Room {random.randint(1, 10)}',
                        'facilitator': 'TBA',
                        'color': colors[j % len(colors)],
                        'duration': 2,  # 2 hour session
                    }
        
        return timetable


# =============================================================================
# PERFORMANCE DASHBOARD
# =============================================================================

class PerformanceDashboardView(LoginRequiredMixin, TemplateView):
    """
    Performance dashboard showing learner progress and at-risk students
    """
    template_name = 'learners/performance.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get filter parameters
        qualification_id = self.request.GET.get('qualification')
        campus_id = self.request.GET.get('campus')
        
        # Base querysets
        enrollments = Enrollment.objects.filter(status__in=['ACTIVE', 'ENROLLED'])
        if qualification_id:
            enrollments = enrollments.filter(qualification_id=qualification_id)
        if campus_id:
            enrollments = enrollments.filter(campus_id=campus_id)
        
        # Overall statistics
        total_active = enrollments.count()
        
        # Get assessment results for these enrollments
        assessment_results = AssessmentResult.objects.filter(
            enrollment__in=enrollments
        )
        
        # Calculate pass rates
        total_assessments = assessment_results.count()
        competent_count = assessment_results.filter(result='C').count()
        nyc_count = assessment_results.filter(result='NYC').count()
        
        pass_rate = (competent_count / total_assessments * 100) if total_assessments > 0 else 0
        
        # Identify at-risk learners (2+ NYC results)
        at_risk_learners = enrollments.annotate(
            nyc_count=Count('assessmentresult', filter=Q(assessmentresult__result='NYC'))
        ).filter(nyc_count__gte=2).select_related('learner', 'qualification')[:20]
        
        # Top performers
        top_performers = enrollments.annotate(
            competent_count=Count('assessmentresult', filter=Q(assessmentresult__result='C')),
            total_assessments=Count('assessmentresult'),
            avg_score=Avg('assessmentresult__percentage_score')
        ).filter(
            total_assessments__gte=3
        ).order_by('-competent_count', '-avg_score').select_related('learner', 'qualification')[:10]
        
        # Performance by qualification
        qual_performance = Enrollment.objects.filter(
            status__in=['ACTIVE', 'ENROLLED']
        ).values(
            'qualification__short_title'
        ).annotate(
            total=Count('id'),
            competent=Count('assessmentresult', filter=Q(assessmentresult__result='C')),
            nyc=Count('assessmentresult', filter=Q(assessmentresult__result='NYC')),
        ).order_by('qualification__short_title')
        
        # Calculate pass rates for each qualification
        for qual in qual_performance:
            total = qual['competent'] + qual['nyc']
            qual['pass_rate'] = round((qual['competent'] / total * 100) if total > 0 else 0, 1)
        
        # Recent assessment activity
        recent_results = assessment_results.select_related(
            'enrollment__learner', 'activity'
        ).order_by('-assessment_date')[:20]
        
        context.update({
            'total_active': total_active,
            'total_assessments': total_assessments,
            'competent_count': competent_count,
            'nyc_count': nyc_count,
            'pass_rate': round(pass_rate, 1),
            'at_risk_learners': at_risk_learners,
            'top_performers': top_performers,
            'qual_performance': qual_performance,
            'recent_results': recent_results,
            'qualifications': Qualification.objects.filter(is_active=True),
            'campuses': Campus.objects.filter(is_active=True),
            'selected_qualification': qualification_id,
            'selected_campus': campus_id,
        })
        
        return context


# =============================================================================
# BULK IMPORT
# =============================================================================

class BulkImportView(LoginRequiredMixin, TemplateView):
    """
    Bulk import view for uploading learner data via CSV/Excel
    """
    template_name = 'learners/bulk_import.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Define import templates
        import_templates = [
            {
                'name': 'Learner Import',
                'description': 'Import new learners with basic information',
                'fields': ['sa_id_number', 'first_name', 'last_name', 'email', 'phone_mobile', 'date_of_birth', 'gender', 'population_group'],
                'template_file': 'learner_import_template.csv'
            },
            {
                'name': 'Enrollment Import',
                'description': 'Import enrollments for existing learners',
                'fields': ['sa_id_number', 'qualification_saqa_id', 'campus_code', 'enrollment_date', 'funding_type'],
                'template_file': 'enrollment_import_template.csv'
            },
            {
                'name': 'Assessment Results Import',
                'description': 'Import assessment results in bulk',
                'fields': ['enrollment_number', 'module_code', 'result', 'assessment_date', 'percentage_score'],
                'template_file': 'assessment_import_template.csv'
            },
        ]
        
        context.update({
            'import_templates': import_templates,
            'qualifications': Qualification.objects.filter(is_active=True),
            'campuses': Campus.objects.filter(is_active=True),
        })
        
        return context
    
    def post(self, request, *args, **kwargs):
        """Handle file upload and import"""
        import_type = request.POST.get('import_type')
        file = request.FILES.get('file')
        
        if not file:
            messages.error(request, 'Please select a file to upload')
            return redirect('learners:bulk_import')
        
        try:
            # Read CSV file
            decoded_file = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded_file))
            
            results = {
                'success': 0,
                'errors': [],
                'warnings': []
            }
            
            if import_type == 'learners':
                results = self._import_learners(reader)
            elif import_type == 'enrollments':
                results = self._import_enrollments(reader)
            elif import_type == 'assessments':
                results = self._import_assessments(reader)
            
            if results['success'] > 0:
                messages.success(request, f"Successfully imported {results['success']} records")
            
            if results['errors']:
                for error in results['errors'][:10]:  # Show first 10 errors
                    messages.error(request, error)
            
            if results['warnings']:
                for warning in results['warnings'][:5]:
                    messages.warning(request, warning)
                    
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
        
        return redirect('learners:bulk_import')
    
    def _import_learners(self, reader):
        """Import learners from CSV"""
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        results = {'success': 0, 'errors': [], 'warnings': []}
        
        for row_num, row in enumerate(reader, start=2):
            try:
                # Check if learner exists
                sa_id = row.get('sa_id_number', '').strip()
                if not sa_id:
                    results['errors'].append(f'Row {row_num}: SA ID number is required')
                    continue
                
                if Learner.objects.filter(sa_id_number=sa_id).exists():
                    results['warnings'].append(f'Row {row_num}: Learner {sa_id} already exists, skipping')
                    continue
                
                # Create user
                email = row.get('email', '').strip()
                if not email:
                    email = f"{sa_id}@placeholder.skillsflow.co.za"
                
                user, created = User.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': row.get('first_name', '').strip(),
                        'last_name': row.get('last_name', '').strip(),
                    }
                )
                if created:
                    user.set_password('changeme123')
                    user.save()
                
                # Create address
                address = Address.objects.create(
                    line_1=row.get('address_line1', 'Not provided'),
                    city=row.get('city', 'Not provided'),
                    province=row.get('province', 'Not provided'),
                    postal_code=row.get('postal_code', '0000'),
                )
                
                # Parse date of birth
                dob_str = row.get('date_of_birth', '').strip()
                try:
                    dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
                except:
                    dob = date(2000, 1, 1)
                
                # Create learner
                learner = Learner.objects.create(
                    user=user,
                    learner_number=f"SKF{timezone.now().year}{row_num:05d}",
                    sa_id_number=sa_id,
                    first_name=row.get('first_name', '').strip(),
                    last_name=row.get('last_name', '').strip(),
                    email=email,
                    phone_mobile=row.get('phone_mobile', '').strip(),
                    date_of_birth=dob,
                    gender=row.get('gender', 'M').strip()[:1].upper(),
                    population_group=row.get('population_group', 'A').strip()[:1].upper(),
                    citizenship='SA',
                    physical_address=address,
                    popia_consent_given=True,
                    popia_consent_date=timezone.now(),
                )
                
                results['success'] += 1
                
            except Exception as e:
                results['errors'].append(f'Row {row_num}: {str(e)}')
        
        return results
    
    def _import_enrollments(self, reader):
        """Import enrollments from CSV"""
        results = {'success': 0, 'errors': [], 'warnings': []}
        
        for row_num, row in enumerate(reader, start=2):
            try:
                sa_id = row.get('sa_id_number', '').strip()
                saqa_id = row.get('qualification_saqa_id', '').strip()
                
                learner = Learner.objects.filter(sa_id_number=sa_id).first()
                if not learner:
                    results['errors'].append(f'Row {row_num}: Learner {sa_id} not found')
                    continue
                
                qualification = Qualification.objects.filter(saqa_id=saqa_id).first()
                if not qualification:
                    results['errors'].append(f'Row {row_num}: Qualification {saqa_id} not found')
                    continue
                
                # Check for existing enrollment
                if Enrollment.objects.filter(learner=learner, qualification=qualification).exists():
                    results['warnings'].append(f'Row {row_num}: Enrollment already exists for {sa_id} in {saqa_id}')
                    continue
                
                # Parse enrollment date
                enroll_date_str = row.get('enrollment_date', '').strip()
                try:
                    enroll_date = datetime.strptime(enroll_date_str, '%Y-%m-%d').date()
                except:
                    enroll_date = date.today()
                
                # Get campus
                campus_code = row.get('campus_code', '').strip()
                campus = Campus.objects.filter(code=campus_code).first() or Campus.objects.first()
                
                # Create enrollment
                enrollment = Enrollment.objects.create(
                    learner=learner,
                    qualification=qualification,
                    campus=campus,
                    enrollment_number=f"ENR{timezone.now().year}{row_num:05d}",
                    enrollment_date=enroll_date,
                    start_date=enroll_date,
                    expected_completion=enroll_date + timedelta(days=qualification.minimum_duration_months * 30),
                    status='ENROLLED',
                    funding_type=row.get('funding_type', 'SELF').strip().upper(),
                )
                
                results['success'] += 1
                
            except Exception as e:
                results['errors'].append(f'Row {row_num}: {str(e)}')
        
        return results
    
    def _import_assessments(self, reader):
        """Import assessment results from CSV"""
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        results = {'success': 0, 'errors': [], 'warnings': []}
        
        # Get a default assessor
        assessor = User.objects.filter(is_staff=True).first()
        
        for row_num, row in enumerate(reader, start=2):
            try:
                enrollment_number = row.get('enrollment_number', '').strip()
                module_code = row.get('module_code', '').strip()
                
                enrollment = Enrollment.objects.filter(enrollment_number=enrollment_number).first()
                if not enrollment:
                    results['errors'].append(f'Row {row_num}: Enrollment {enrollment_number} not found')
                    continue
                
                # Find activity by module code
                activity = AssessmentActivity.objects.filter(module__code=module_code).first()
                if not activity:
                    results['errors'].append(f'Row {row_num}: Activity for module {module_code} not found')
                    continue
                
                # Parse assessment date
                assess_date_str = row.get('assessment_date', '').strip()
                try:
                    assess_date = datetime.strptime(assess_date_str, '%Y-%m-%d').date()
                except:
                    assess_date = date.today()
                
                # Create result
                result_code = row.get('result', 'C').strip().upper()
                if result_code not in ['C', 'NYC', 'ABS', 'DEF']:
                    result_code = 'C'
                
                AssessmentResult.objects.create(
                    enrollment=enrollment,
                    activity=activity,
                    assessor=assessor,
                    result=result_code,
                    percentage_score=Decimal(row.get('percentage_score', '0') or '0'),
                    assessment_date=assess_date,
                    status='MODERATED' if result_code == 'C' else 'PENDING_MOD',
                )
                
                results['success'] += 1
                
            except Exception as e:
                results['errors'].append(f'Row {row_num}: {str(e)}')
        
        return results


@login_required
def download_import_template(request, template_type):
    """Download CSV template for imports"""
    templates = {
        'learners': {
            'filename': 'learner_import_template.csv',
            'headers': ['sa_id_number', 'first_name', 'last_name', 'email', 'phone_mobile', 
                       'date_of_birth', 'gender', 'population_group', 'address_line1', 
                       'city', 'province', 'postal_code'],
            'sample': ['9001015800086', 'John', 'Doe', 'john@email.com', '0821234567',
                      '1990-01-01', 'M', 'A', '123 Main Street', 'Johannesburg', 'Gauteng', '2000']
        },
        'enrollments': {
            'filename': 'enrollment_import_template.csv',
            'headers': ['sa_id_number', 'qualification_saqa_id', 'campus_code', 
                       'enrollment_date', 'funding_type'],
            'sample': ['9001015800086', 'SAQA-12345', 'SKF-JHB', '2025-01-15', 'SELF']
        },
        'assessments': {
            'filename': 'assessment_import_template.csv',
            'headers': ['enrollment_number', 'module_code', 'result', 
                       'assessment_date', 'percentage_score'],
            'sample': ['ENR202500001', '12345-M01', 'C', '2025-06-15', '75.5']
        }
    }
    
    template = templates.get(template_type)
    if not template:
        return HttpResponse('Template not found', status=404)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{template["filename"]}"'
    
    writer = csv.writer(response)
    writer.writerow(template['headers'])
    writer.writerow(template['sample'])
    
    return response


# =============================================================================
# SETA EXPORT TEMPLATES
# =============================================================================

class ExportTemplatesView(LoginRequiredMixin, TemplateView):
    """
    Manage and use export templates for SETA submissions
    """
    template_name = 'learners/export_templates.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Predefined SETA export templates
        export_templates = [
            {
                'id': 'nlrd_learner',
                'name': 'NLRD Learner Upload',
                'description': 'National Learners Records Database - Learner information submission',
                'seta': 'All SETAs',
                'fields': 20,
                'format': 'CSV',
            },
            {
                'id': 'nlrd_enrollment',
                'name': 'NLRD Enrollment Upload',
                'description': 'Learner enrollment/registration data for NLRD',
                'seta': 'All SETAs',
                'fields': 25,
                'format': 'CSV',
            },
            {
                'id': 'nlrd_achievement',
                'name': 'NLRD Achievement Upload',
                'description': 'Learner achievement/completion records',
                'seta': 'All SETAs',
                'fields': 15,
                'format': 'CSV',
            },
            {
                'id': 'wsp_training_report',
                'name': 'WSP Training Report',
                'description': 'Workplace Skills Plan training implementation report',
                'seta': 'Services SETA',
                'fields': 30,
                'format': 'Excel',
            },
            {
                'id': 'atr_summary',
                'name': 'ATR Summary Report',
                'description': 'Annual Training Report summary for SETA submission',
                'seta': 'All SETAs',
                'fields': 18,
                'format': 'Excel',
            },
            {
                'id': 'qcto_assessment',
                'name': 'QCTO Assessment Records',
                'description': 'Quality Council for Trades and Occupations assessment records',
                'seta': 'QCTO',
                'fields': 22,
                'format': 'CSV',
            },
        ]
        
        context.update({
            'export_templates': export_templates,
            'qualifications': Qualification.objects.filter(is_active=True),
            'campuses': Campus.objects.filter(is_active=True),
            'setas': SETA.objects.filter(is_active=True),
        })
        
        return context


@login_required
def export_seta_data(request, template_id):
    """Generate SETA export based on template"""
    # Get filter parameters
    qualification_id = request.GET.get('qualification')
    campus_id = request.GET.get('campus')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Base queryset
    enrollments = Enrollment.objects.select_related(
        'learner', 'qualification', 'campus', 'learner__physical_address'
    )
    
    if qualification_id:
        enrollments = enrollments.filter(qualification_id=qualification_id)
    if campus_id:
        enrollments = enrollments.filter(campus_id=campus_id)
    if date_from:
        enrollments = enrollments.filter(enrollment_date__gte=date_from)
    if date_to:
        enrollments = enrollments.filter(enrollment_date__lte=date_to)
    
    # Generate appropriate export
    if template_id == 'nlrd_learner':
        return _export_nlrd_learner(enrollments)
    elif template_id == 'nlrd_enrollment':
        return _export_nlrd_enrollment(enrollments)
    elif template_id == 'nlrd_achievement':
        return _export_nlrd_achievement(enrollments)
    elif template_id == 'wsp_training_report':
        return _export_wsp_report(enrollments)
    elif template_id == 'atr_summary':
        return _export_atr_summary(enrollments)
    elif template_id == 'qcto_assessment':
        return _export_qcto_assessment(enrollments)
    
    return HttpResponse('Template not found', status=404)


def _export_nlrd_learner(enrollments):
    """Export NLRD Learner Upload format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="NLRD_Learner_{date.today()}.csv"'
    
    writer = csv.writer(response)
    
    # NLRD Learner file header
    writer.writerow([
        'Natl_ID', 'Person_Alternate_ID', 'Alternate_ID_Type', 'Equity_Code',
        'Nationality_Code', 'Home_Language_Code', 'Gender', 'Citizen_Resident_Status_Code',
        'Socio_Economic_Status_Code', 'Disability_Status_Code', 'First_Name',
        'Second_Name', 'Surname', 'Birth_Date', 'School_EMIS_Number',
        'Highest_Edu_Level_Code', 'Physical_Address_Line_1', 'Physical_Address_Line_2',
        'Physical_Address_City', 'Physical_Address_Postcode'
    ])
    
    # Export unique learners
    exported_ids = set()
    for enrollment in enrollments:
        learner = enrollment.learner
        if learner.sa_id_number in exported_ids:
            continue
        exported_ids.add(learner.sa_id_number)
        
        address = learner.physical_address
        
        writer.writerow([
            learner.sa_id_number,
            '',  # Alternate ID
            '',  # Alternate ID Type
            learner.population_group,
            'SA' if learner.citizenship == 'SA' else learner.passport_country,
            learner.home_language[:3] if learner.home_language else 'ENG',
            learner.gender,
            'SA' if learner.citizenship == 'SA' else 'O',
            learner.socio_economic_status or 'U',
            learner.disability_status,
            learner.first_name,
            learner.middle_name or '',
            learner.last_name,
            learner.date_of_birth.strftime('%Y%m%d') if learner.date_of_birth else '',
            '',  # School EMIS
            learner.highest_qualification or '4',
            address.line_1 if address else '',
            address.line_2 if address else '',
            address.city if address else '',
            address.postal_code if address else '',
        ])
    
    return response


def _export_nlrd_enrollment(enrollments):
    """Export NLRD Enrollment Upload format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="NLRD_Enrollment_{date.today()}.csv"'
    
    writer = csv.writer(response)
    
    # NLRD Enrollment header
    writer.writerow([
        'Natl_ID', 'Person_Alternate_ID', 'Alternate_ID_Type', 'Enrol_Status_Code',
        'Enrol_Status_Start_Date', 'Enrol_Date', 'Provider_Code', 'Provider_ETQI_ID',
        'Qualification_ID', 'Learnership_ID', 'Funding_Type', 'Cumulative_Spend',
        'OFO_Code', 'Urban_Rural_Code', 'SDL_Number', 'Site_Number',
        'Practical_Provider_Code', 'Practical_Provider_ETQE_ID', 'SIC_Code',
        'Non_NFSD_Funding_Source', 'Assessment_ETQE_ID', 'Enrolment_NQF_Level',
        'Part_of_ID', 'Last_School_Year', 'Last_School_EMIS_Number'
    ])
    
    for enrollment in enrollments:
        learner = enrollment.learner
        qual = enrollment.qualification
        
        # Map status codes
        status_map = {
            'APPLIED': '1', 'DOC_CHECK': '1', 'REGISTERED': '2', 'ENROLLED': '2',
            'ACTIVE': '3', 'ON_HOLD': '5', 'COMPLETED': '4', 'CERTIFIED': '4',
            'WITHDRAWN': '6', 'TRANSFERRED': '7', 'EXPIRED': '8'
        }
        
        writer.writerow([
            learner.sa_id_number,
            '',
            '',
            status_map.get(enrollment.status, '3'),
            enrollment.enrollment_date.strftime('%Y%m%d') if enrollment.enrollment_date else '',
            enrollment.enrollment_date.strftime('%Y%m%d') if enrollment.enrollment_date else '',
            enrollment.campus.code if enrollment.campus else '',
            '',  # Provider ETQI
            qual.saqa_id if qual else '',
            '',  # Learnership ID
            enrollment.funding_type,
            '',  # Cumulative Spend
            '',  # OFO Code
            '',  # Urban/Rural
            '',  # SDL Number
            '',  # Site Number
            '',  # Practical Provider
            '',  # Practical Provider ETQE
            '',  # SIC Code
            '',  # Non-NFSD Funding
            '',  # Assessment ETQE
            qual.nqf_level if qual else '',
            '',  # Part of ID
            '',  # Last School Year
            '',  # Last School EMIS
        ])
    
    return response


def _export_nlrd_achievement(enrollments):
    """Export NLRD Achievement Upload format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="NLRD_Achievement_{date.today()}.csv"'
    
    writer = csv.writer(response)
    
    # NLRD Achievement header
    writer.writerow([
        'Natl_ID', 'Person_Alternate_ID', 'Alternate_ID_Type', 'Qualification_ID',
        'Designation_Code', 'Designation_Start_Date', 'Designation_End_Date',
        'Provider_Code', 'Provider_ETQI_ID', 'Assessment_ETQE_ID',
        'NLRD_Certificate_Number', 'Certificate_Number', 'Honour_Code',
        'Part_Of_ID', 'Achievement_Date'
    ])
    
    # Export completed/certified enrollments
    for enrollment in enrollments.filter(status__in=['COMPLETED', 'CERTIFIED']):
        learner = enrollment.learner
        qual = enrollment.qualification
        
        writer.writerow([
            learner.sa_id_number,
            '',
            '',
            qual.saqa_id if qual else '',
            'Q',  # Qualification
            enrollment.enrollment_date.strftime('%Y%m%d') if enrollment.enrollment_date else '',
            enrollment.actual_completion.strftime('%Y%m%d') if enrollment.actual_completion else '',
            enrollment.campus.code if enrollment.campus else '',
            '',
            '',
            enrollment.nlrd_reference or '',
            enrollment.certificate_number or '',
            '',  # Honour code
            '',
            enrollment.actual_completion.strftime('%Y%m%d') if enrollment.actual_completion else '',
        ])
    
    return response


def _export_wsp_report(enrollments):
    """Export WSP Training Report format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="WSP_Training_Report_{date.today()}.csv"'
    
    writer = csv.writer(response)
    
    writer.writerow([
        'Employee_ID', 'First_Name', 'Last_Name', 'ID_Number', 'Gender',
        'Race', 'Disability', 'Province', 'Occupation', 'OFO_Code',
        'Training_Type', 'Qualification_Name', 'SAQA_ID', 'NQF_Level',
        'Start_Date', 'End_Date', 'Status', 'Provider_Name', 'Provider_Accreditation',
        'Funding_Type', 'Training_Cost', 'Completion_Status'
    ])
    
    for enrollment in enrollments:
        learner = enrollment.learner
        qual = enrollment.qualification
        
        writer.writerow([
            learner.learner_number,
            learner.first_name,
            learner.last_name,
            learner.sa_id_number,
            learner.gender,
            learner.population_group,
            learner.disability_status,
            learner.province_code,
            '',  # Occupation
            '',  # OFO Code
            qual.qualification_type if qual else '',
            qual.title if qual else '',
            qual.saqa_id if qual else '',
            qual.nqf_level if qual else '',
            enrollment.start_date.strftime('%Y-%m-%d') if enrollment.start_date else '',
            enrollment.expected_completion.strftime('%Y-%m-%d') if enrollment.expected_completion else '',
            enrollment.status,
            enrollment.campus.name if enrollment.campus else '',
            '',  # Accreditation
            enrollment.funding_type,
            '',  # Cost
            'Completed' if enrollment.status in ['COMPLETED', 'CERTIFIED'] else 'In Progress',
        ])
    
    return response


def _export_atr_summary(enrollments):
    """Export ATR Summary format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="ATR_Summary_{date.today()}.csv"'
    
    writer = csv.writer(response)
    
    # Summary header
    writer.writerow(['Metric', 'Value'])
    
    # Calculate summary stats
    total = enrollments.count()
    completed = enrollments.filter(status__in=['COMPLETED', 'CERTIFIED']).count()
    active = enrollments.filter(status='ACTIVE').count()
    
    # Gender breakdown
    male = enrollments.filter(learner__gender='M').count()
    female = enrollments.filter(learner__gender='F').count()
    
    # Population group breakdown  
    african = enrollments.filter(learner__population_group='A').count()
    coloured = enrollments.filter(learner__population_group='C').count()
    indian = enrollments.filter(learner__population_group='I').count()
    white = enrollments.filter(learner__population_group='W').count()
    
    writer.writerow(['Total Enrollments', total])
    writer.writerow(['Completed', completed])
    writer.writerow(['Active', active])
    writer.writerow(['Completion Rate %', round(completed/total*100, 1) if total > 0 else 0])
    writer.writerow(['', ''])
    writer.writerow(['Gender Breakdown', ''])
    writer.writerow(['Male', male])
    writer.writerow(['Female', female])
    writer.writerow(['', ''])
    writer.writerow(['Population Group', ''])
    writer.writerow(['African', african])
    writer.writerow(['Coloured', coloured])
    writer.writerow(['Indian', indian])
    writer.writerow(['White', white])
    
    return response


def _export_qcto_assessment(enrollments):
    """Export QCTO Assessment Records format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="QCTO_Assessment_{date.today()}.csv"'
    
    writer = csv.writer(response)
    
    writer.writerow([
        'Learner_ID', 'First_Name', 'Last_Name', 'ID_Number',
        'Qualification_Code', 'Module_Code', 'Module_Title',
        'Assessment_Type', 'Assessor_ID', 'Assessment_Date',
        'Result', 'Score', 'Attempt_Number', 'Comments'
    ])
    
    # Get assessment results for these enrollments
    results = AssessmentResult.objects.filter(
        enrollment__in=enrollments
    ).select_related('enrollment__learner', 'activity__module', 'assessor')
    
    for result in results:
        learner = result.enrollment.learner
        activity = result.activity
        module = activity.module if activity else None
        
        writer.writerow([
            learner.learner_number,
            learner.first_name,
            learner.last_name,
            learner.sa_id_number,
            result.enrollment.qualification.saqa_id if result.enrollment.qualification else '',
            module.code if module else '',
            module.title if module else '',
            activity.activity_type if activity else '',
            result.assessor.email if result.assessor else '',
            result.assessment_date.strftime('%Y-%m-%d') if result.assessment_date else '',
            result.result,
            str(result.percentage_score) if result.percentage_score else '',
            result.attempt_number,
            result.feedback or '',
        ])
    
    return response


# =============================================================================
# LEARNER LIST & DETAIL VIEWS
# =============================================================================

class LearnerListView(LoginRequiredMixin, ListView):
    """
    Standard list view for learners via their enrollments.
    Shows learner + qualification combinations (enrollments).
    """
    model = Enrollment
    template_name = 'learners/list.html'
    context_object_name = 'enrollments'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = Enrollment.objects.select_related(
            'learner', 'learner__campus', 'learner__physical_address',
            'qualification', 'campus'
        ).order_by('-enrollment_date')
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(learner__first_name__icontains=search) |
                Q(learner__last_name__icontains=search) |
                Q(learner__sa_id_number__icontains=search) |
                Q(learner__email__icontains=search) |
                Q(learner__learner_number__icontains=search) |
                Q(enrollment_number__icontains=search)
            )
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by qualification
        qualification_id = self.request.GET.get('qualification')
        if qualification_id:
            queryset = queryset.filter(qualification_id=qualification_id)
        
        # Filter by campus - use URL param if provided, else use global campus filter
        campus_id = self.request.GET.get('campus')
        if campus_id:
            queryset = queryset.filter(campus_id=campus_id)
        else:
            # Apply global campus filter if set
            selected_campus = get_selected_campus(self.request)
            if selected_campus:
                queryset = queryset.filter(campus=selected_campus)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['campuses'] = Campus.objects.filter(is_active=True)
        context['qualifications'] = Qualification.objects.filter(is_active=True)
        context['search'] = self.request.GET.get('search', '')
        context['selected_campus'] = self.request.GET.get('campus', '')
        context['selected_qualification'] = self.request.GET.get('qualification', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['total_count'] = self.get_queryset().count()
        return context


class LearnerDetailView(LoginRequiredMixin, DetailView):
    """Detailed learner profile view"""
    model = Learner
    template_name = 'learners/detail.html'
    context_object_name = 'learner'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        learner = self.object
        
        # Get enrollments
        enrollments = Enrollment.objects.filter(learner=learner).select_related(
            'qualification', 'campus'
        ).order_by('-enrollment_date')
        
        # Get assessment results
        results = AssessmentResult.objects.filter(
            enrollment__learner=learner
        ).select_related('activity', 'enrollment__qualification').order_by('-assessment_date')
        
        # Calculate stats
        total_assessments = results.count()
        competent_count = results.filter(result='C').count()
        pass_rate = (competent_count / total_assessments * 100) if total_assessments > 0 else 0
        
        context.update({
            'enrollments': enrollments,
            'recent_results': results[:10],
            'total_assessments': total_assessments,
            'competent_count': competent_count,
            'pass_rate': round(pass_rate, 1),
        })
        
        return context


# =============================================================================
# DAILY LOGBOOK & MONTHLY EXPORT
# =============================================================================

@login_required
def monthly_logbook_export(request, placement_id, year, month):
    """
    Export monthly logbook entries to Excel format.
    Aggregates daily entries with task completions.
    """
    from django.http import HttpResponse
    from corporate.models import WorkplacePlacement
    from .models import DailyLogbookEntry, DailyTaskCompletion
    from calendar import monthrange, month_name
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "Excel export requires openpyxl. Please install it: pip install openpyxl",
            status=500
        )
    
    # Get placement
    placement = get_object_or_404(WorkplacePlacement, pk=placement_id)
    
    # Security check - ensure user can access this placement
    if not request.user.is_staff:
        if hasattr(request.user, 'learner_profile'):
            if placement.learner != request.user.learner_profile:
                return HttpResponse("Unauthorized", status=403)
        else:
            return HttpResponse("Unauthorized", status=403)
    
    # Get entries for the month
    entries = DailyLogbookEntry.objects.filter(
        placement=placement,
        entry_date__year=year,
        entry_date__month=month
    ).prefetch_related('task_completions__workplace_outcome').order_by('entry_date')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Logbook {month_name[month]} {year}"
    
    # Styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    alt_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    
    # Title section
    ws.merge_cells('A1:H1')
    ws['A1'] = f"WORKPLACE LOGBOOK - {month_name[month].upper()} {year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    
    # Learner info
    learner = placement.learner
    ws['A3'] = "Learner Name:"
    ws['B3'] = learner.get_full_name()
    ws['A4'] = "ID Number:"
    ws['B4'] = learner.id_number
    ws['A5'] = "Qualification:"
    ws['B5'] = placement.enrollment.qualification.short_title if placement.enrollment else "N/A"
    ws['D3'] = "Host Employer:"
    ws['E3'] = placement.host_employer.name if placement.host_employer else "N/A"
    ws['D4'] = "Mentor:"
    ws['E4'] = placement.mentor.user.get_full_name() if placement.mentor else "N/A"
    ws['D5'] = "Period:"
    ws['E5'] = f"{month_name[month]} {year}"
    
    # Bold labels
    for cell in ['A3', 'A4', 'A5', 'D3', 'D4', 'D5']:
        ws[cell].font = Font(bold=True)
    
    # Summary section
    summary = DailyLogbookEntry.get_monthly_summary(placement, year, month)
    
    ws['A7'] = "MONTHLY SUMMARY"
    ws['A7'].font = header_font
    ws.merge_cells('A7:H7')
    
    ws['A8'] = "Days Present:"
    ws['B8'] = summary.get('present_days', 0)
    ws['C8'] = "Days Absent:"
    ws['D8'] = summary.get('absent_days', 0)
    ws['E8'] = "Sick Leave:"
    ws['F8'] = summary.get('sick_days', 0)
    ws['G8'] = "Total Hours:"
    ws['H8'] = round(summary.get('total_hours', 0), 1)
    
    for cell in ['A8', 'C8', 'E8', 'G8']:
        ws[cell].font = Font(bold=True)
    
    # Daily entries header
    header_row = 11
    headers = ['Date', 'Day', 'Status', 'Clock In', 'Clock Out', 'Hours', 'Tasks Completed', 'Daily Summary']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 50
    
    # Daily entries
    row = header_row + 1
    for entry in entries:
        # Get tasks for this day
        tasks = entry.task_completions.all()
        task_summary = ", ".join([
            f"{t.outcome_code}: {t.task_description[:30]}..." 
            for t in tasks[:5]
        ])
        if tasks.count() > 5:
            task_summary += f" (+{tasks.count() - 5} more)"
        
        ws.cell(row=row, column=1, value=entry.entry_date.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row, column=2, value=entry.entry_date.strftime('%a')).border = thin_border
        ws.cell(row=row, column=3, value=entry.get_attendance_status_display()).border = thin_border
        ws.cell(row=row, column=4, value=entry.clock_in.strftime('%H:%M') if entry.clock_in else '-').border = thin_border
        ws.cell(row=row, column=5, value=entry.clock_out.strftime('%H:%M') if entry.clock_out else '-').border = thin_border
        ws.cell(row=row, column=6, value=entry.hours_worked).border = thin_border
        ws.cell(row=row, column=7, value=task_summary).border = thin_border
        ws.cell(row=row, column=8, value=entry.daily_summary).border = thin_border
        
        # Wrap text for longer fields
        ws.cell(row=row, column=7).alignment = wrap_align
        ws.cell(row=row, column=8).alignment = wrap_align
        
        # Alternate row coloring
        if row % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = alt_fill
        
        row += 1
    
    # Tasks detail sheet
    ws2 = wb.create_sheet(title="Task Details")
    
    # Tasks header
    task_headers = ['Date', 'Outcome Code', 'Module', 'Task Description', 'Hours', 'Competency', 'Evidence Notes']
    for col, header in enumerate(task_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Task column widths
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 60
    ws2.column_dimensions['E'].width = 8
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 40
    
    # All tasks for the month
    all_tasks = DailyTaskCompletion.objects.filter(
        daily_entry__placement=placement,
        daily_entry__entry_date__year=year,
        daily_entry__entry_date__month=month
    ).select_related('daily_entry', 'workplace_outcome__module').order_by('daily_entry__entry_date')
    
    task_row = 2
    for task in all_tasks:
        ws2.cell(row=task_row, column=1, value=task.daily_entry.entry_date.strftime('%Y-%m-%d')).border = thin_border
        ws2.cell(row=task_row, column=2, value=task.outcome_code).border = thin_border
        ws2.cell(row=task_row, column=3, value=task.module_code).border = thin_border
        ws2.cell(row=task_row, column=4, value=task.task_description).border = thin_border
        ws2.cell(row=task_row, column=5, value=float(task.hours_spent)).border = thin_border
        ws2.cell(row=task_row, column=6, value=task.get_competency_rating_display() if task.competency_rating else '-').border = thin_border
        ws2.cell(row=task_row, column=7, value=task.evidence_notes).border = thin_border
        
        ws2.cell(row=task_row, column=4).alignment = wrap_align
        ws2.cell(row=task_row, column=7).alignment = wrap_align
        
        if task_row % 2 == 0:
            for col in range(1, 8):
                ws2.cell(row=task_row, column=col).fill = alt_fill
        
        task_row += 1
    
    # Sign-off section on main sheet
    sign_row = row + 2
    ws.cell(row=sign_row, column=1, value="SIGN-OFF").font = header_font
    ws.merge_cells(f'A{sign_row}:H{sign_row}')
    
    sign_row += 2
    ws.cell(row=sign_row, column=1, value="Learner Signature:")
    ws.cell(row=sign_row, column=2, value="_____________________")
    ws.cell(row=sign_row, column=4, value="Date:")
    ws.cell(row=sign_row, column=5, value="_____________________")
    
    sign_row += 2
    ws.cell(row=sign_row, column=1, value="Mentor Signature:")
    ws.cell(row=sign_row, column=2, value="_____________________")
    ws.cell(row=sign_row, column=4, value="Date:")
    ws.cell(row=sign_row, column=5, value="_____________________")
    
    sign_row += 2
    ws.cell(row=sign_row, column=1, value="Facilitator Signature:")
    ws.cell(row=sign_row, column=2, value="_____________________")
    ws.cell(row=sign_row, column=4, value="Date:")
    ws.cell(row=sign_row, column=5, value="_____________________")
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Logbook_{learner.get_full_name().replace(' ', '_')}_{month_name[month]}_{year}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def daily_logbook_calendar_view(request, placement_id):
    """
    Calendar view for daily logbook entries.
    Shows month at a time with attendance status coloring.
    """
    from corporate.models import WorkplacePlacement
    from .models import DailyLogbookEntry
    from calendar import monthcalendar, month_name
    
    placement = get_object_or_404(WorkplacePlacement, pk=placement_id)
    
    # Security check
    if not request.user.is_staff:
        if hasattr(request.user, 'learner_profile'):
            if placement.learner != request.user.learner_profile:
                return HttpResponse("Unauthorized", status=403)
        else:
            return HttpResponse("Unauthorized", status=403)
    
    # Get year/month from query params, default to current
    year = int(request.GET.get('year', date.today().year))
    month = int(request.GET.get('month', date.today().month))
    
    # Get entries for the month
    entries = DailyLogbookEntry.objects.filter(
        placement=placement,
        entry_date__year=year,
        entry_date__month=month
    ).prefetch_related('task_completions')
    
    # Create lookup dict
    entries_by_date = {e.entry_date.day: e for e in entries}
    
    # Build calendar data
    cal = monthcalendar(year, month)
    calendar_weeks = []
    
    for week in cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append({'day': None, 'entry': None})
            else:
                entry = entries_by_date.get(day)
                week_data.append({
                    'day': day,
                    'date': date(year, month, day),
                    'entry': entry,
                    'status': entry.attendance_status if entry else None,
                    'tasks_count': entry.tasks_count if entry else 0,
                    'hours': entry.hours_worked if entry else 0,
                    'is_weekend': date(year, month, day).weekday() >= 5,
                })
        calendar_weeks.append(week_data)
    
    # Navigation
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1
    
    # Get summary
    summary = DailyLogbookEntry.get_monthly_summary(placement, year, month)
    
    context = {
        'placement': placement,
        'learner': placement.learner,
        'year': year,
        'month': month,
        'month_name': month_name[month],
        'calendar_weeks': calendar_weeks,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'summary': summary,
        'entries': entries,
    }
    
    return render(request, 'learners/logbook_calendar.html', context)

