"""
Attendance Register Service

Generates monthly attendance registers for Training Notification projects.
Aggregates learner attendance data, verification status, and stipend calculations.
Supports PDF and Excel export formats.
"""
import calendar
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Any

from django.db.models import Count, Q, Sum, F
from django.template.loader import render_to_string
from django.utils import timezone
from django.core.files.base import ContentFile

# Optional imports for PDF/Excel generation
try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class AttendanceRegisterService:
    """
    Service to generate monthly attendance registers for NOT projects.
    
    Usage:
        service = AttendanceRegisterService(training_notification, year, month)
        data = service.get_register_data()
        pdf_content = service.generate_pdf()
        excel_content = service.generate_excel()
    """
    
    # Attendance type display mappings
    ATTENDANCE_CODES = {
        'PRESENT': 'P',
        'ANNUAL': 'AL',
        'SICK': 'SL',
        'FAMILY': 'FL',
        'UNPAID': 'UL',
        'PUBLIC_HOLIDAY': 'PH',
        'ABSENT': 'A',
        'SUSPENDED': 'S',
    }
    
    ATTENDANCE_COLORS = {
        'PRESENT': '#22c55e',  # green
        'ANNUAL': '#3b82f6',   # blue
        'SICK': '#f59e0b',     # amber
        'FAMILY': '#8b5cf6',   # purple
        'UNPAID': '#ef4444',   # red
        'PUBLIC_HOLIDAY': '#06b6d4',  # cyan
        'ABSENT': '#dc2626',   # dark red
        'SUSPENDED': '#6b7280',  # gray
        'NO_RECORD': '#e5e7eb',  # light gray
    }
    
    def __init__(self, training_notification, year: int, month: int):
        """
        Initialize the service.
        
        Args:
            training_notification: TrainingNotification model instance
            year: Year for the register (e.g., 2026)
            month: Month for the register (1-12)
        """
        self.not_project = training_notification
        self.year = year
        self.month = month
        
        # Calculate working days in month
        self.month_start = date(year, month, 1)
        _, last_day = calendar.monthrange(year, month)
        self.month_end = date(year, month, last_day)
        
        # Get all dates in the month
        self.all_dates = [
            self.month_start + timedelta(days=i)
            for i in range((self.month_end - self.month_start).days + 1)
        ]
        
        # Get working days (exclude weekends)
        self.working_days = [d for d in self.all_dates if d.weekday() < 5]
        self.total_working_days = len(self.working_days)
    
    def get_placements(self):
        """
        Get all active workplace placements linked to this NOT project.
        
        The link chain is:
        TrainingNotification -> workplace_placements (direct FK)
        OR
        TrainingNotification -> intakes -> cohort -> enrollments -> workplace_placements
        """
        from corporate.models import WorkplacePlacement
        
        # Direct link via training_notification FK on WorkplacePlacement
        placements = WorkplacePlacement.objects.filter(
            training_notification=self.not_project,
            status__in=['ACTIVE', 'COMPLETED']
        ).select_related(
            'learner', 'learner__user', 'host', 'mentor', 'mentor__user', 'leave_policy'
        ).order_by('learner__last_name', 'learner__first_name')
        
        # If no direct placements, try via intakes -> cohort -> enrollments
        if not placements.exists():
            intake_cohort_ids = self.not_project.intakes.exclude(
                cohort__isnull=True
            ).values_list('cohort_id', flat=True)
            
            placements = WorkplacePlacement.objects.filter(
                enrollment__cohort_id__in=intake_cohort_ids,
                status__in=['ACTIVE', 'COMPLETED']
            ).select_related(
                'learner', 'learner__user', 'host', 'mentor', 'mentor__user', 'leave_policy'
            ).order_by('learner__last_name', 'learner__first_name')
        
        return placements
    
    def get_attendance_for_placement(self, placement) -> Dict[date, Dict]:
        """
        Get all attendance records for a placement in the specified month.
        Returns a dict mapping date -> attendance record data.
        """
        from learners.models import WorkplaceAttendance
        
        records = WorkplaceAttendance.objects.filter(
            placement=placement,
            date__gte=self.month_start,
            date__lte=self.month_end
        ).order_by('date')
        
        attendance_map = {}
        for record in records:
            attendance_map[record.date] = {
                'id': record.id,
                'type': record.attendance_type,
                'code': self.ATTENDANCE_CODES.get(record.attendance_type, '?'),
                'clock_in': record.clock_in,
                'clock_out': record.clock_out,
                'hours_worked': record.hours_worked,
                'mentor_verified': record.mentor_verified,
                'facilitator_verified': record.facilitator_verified,
                'dual_verified': record.mentor_verified and record.facilitator_verified,
                'notes': record.notes,
                'has_photo': bool(record.photo),
                'has_gps': record.gps_latitude is not None,
            }
        
        return attendance_map
    
    def get_stipend_for_placement(self, placement) -> Optional[Dict]:
        """
        Get the stipend calculation for a placement for the specified month.
        """
        from learners.models import StipendCalculation
        
        try:
            stipend = StipendCalculation.objects.get(
                placement=placement,
                year=self.year,
                month=self.month
            )
            return {
                'id': stipend.id,
                'status': stipend.status,
                'total_working_days': stipend.total_working_days,
                'days_present': stipend.days_present,
                'days_annual_leave': stipend.days_annual_leave,
                'days_sick_leave': stipend.days_sick_leave,
                'days_family_leave': stipend.days_family_leave,
                'days_unpaid_leave': stipend.days_unpaid_leave,
                'days_public_holiday': stipend.days_public_holiday,
                'days_absent': stipend.days_absent,
                'days_suspended': stipend.days_suspended,
                'daily_rate': stipend.daily_rate,
                'gross_amount': stipend.gross_amount,
                'total_deductions': stipend.total_deductions,
                'net_amount': stipend.net_amount,
                'dual_verified_records': stipend.dual_verified_records,
                'unverified_records': stipend.unverified_records,
                'verification_percentage': self._calc_verification_pct(stipend),
            }
        except StipendCalculation.DoesNotExist:
            return None
    
    def _calc_verification_pct(self, stipend) -> float:
        """Calculate the percentage of dual-verified records."""
        if stipend.total_attendance_records > 0:
            return round(
                (stipend.dual_verified_records / stipend.total_attendance_records) * 100, 
                1
            )
        return 0.0
    
    def calculate_learner_summary(self, placement, attendance_map: Dict) -> Dict:
        """
        Calculate summary statistics for a learner's attendance.
        """
        # Count attendance types - use defaultdict to handle any type
        from collections import defaultdict
        type_counts = defaultdict(int)
        
        verified_count = 0
        mentor_only = 0
        facilitator_only = 0
        unverified = 0
        total_hours = Decimal('0.00')
        days_with_records = 0
        
        for work_date in self.working_days:
            if work_date in attendance_map:
                record = attendance_map[work_date]
                type_counts[record['type']] += 1
                days_with_records += 1
                
                if record['hours_worked']:
                    total_hours += Decimal(str(record['hours_worked']))
                
                if record['dual_verified']:
                    verified_count += 1
                elif record['mentor_verified']:
                    mentor_only += 1
                elif record['facilitator_verified']:
                    facilitator_only += 1
                else:
                    unverified += 1
        
        # Days with no record (missed logging)
        missed_logging = self.total_working_days - days_with_records
        
        # Calculate payable days based on leave policy
        # Handle various naming conventions for leave types
        payable_days = (
            type_counts['PRESENT'] +
            type_counts['ANNUAL'] + type_counts['ANNUAL_LEAVE'] +
            type_counts['SICK'] + type_counts['SICK_LEAVE'] +
            type_counts['FAMILY'] + type_counts['FAMILY_LEAVE'] +
            type_counts['PUBLIC_HOLIDAY']
        )
        
        # Days that reduce stipend
        deduction_days = (
            type_counts['UNPAID'] + type_counts['UNPAID_LEAVE'] +
            type_counts['ABSENT'] +
            type_counts['SUSPENDED'] +
            missed_logging
        )
        
        # Calculate stipend (if daily rate available)
        daily_rate = placement.stipend_daily_rate or Decimal('0.00')
        gross_amount = payable_days * daily_rate
        deductions = deduction_days * daily_rate
        net_amount = max(Decimal('0.00'), gross_amount - deductions)
        
        # Verification percentage
        if days_with_records > 0:
            verification_pct = round((verified_count / days_with_records) * 100, 1)
        else:
            verification_pct = 0.0
        
        return {
            'type_counts': type_counts,
            'days_with_records': days_with_records,
            'missed_logging': missed_logging,
            'total_hours': total_hours,
            'verification': {
                'dual_verified': verified_count,
                'mentor_only': mentor_only,
                'facilitator_only': facilitator_only,
                'unverified': unverified,
                'percentage': verification_pct,
            },
            'stipend': {
                'daily_rate': daily_rate,
                'payable_days': payable_days,
                'deduction_days': deduction_days,
                'gross_amount': gross_amount,
                'deductions': deductions,
                'net_amount': net_amount,
            }
        }
    
    def get_register_data(self) -> Dict[str, Any]:
        """
        Get complete attendance register data for the NOT project.
        
        Returns:
            Dictionary containing all register data for rendering.
        """
        placements = self.get_placements()
        
        learner_data = []
        totals = {
            'learner_count': 0,
            'total_present': 0,
            'total_absent': 0,
            'total_leave': 0,
            'total_payable_days': 0,
            'total_stipend': Decimal('0.00'),
            'fully_verified': 0,
            'partially_verified': 0,
            'unverified': 0,
        }
        
        for placement in placements:
            attendance_map = self.get_attendance_for_placement(placement)
            summary = self.calculate_learner_summary(placement, attendance_map)
            existing_stipend = self.get_stipend_for_placement(placement)
            
            # Build daily attendance grid
            daily_attendance = []
            for work_date in self.working_days:
                if work_date in attendance_map:
                    record = attendance_map[work_date]
                    daily_attendance.append({
                        'date': work_date,
                        'day': work_date.day,
                        'weekday': work_date.strftime('%a'),
                        'code': record['code'],
                        'type': record['type'],
                        'verified': record['dual_verified'],
                        'mentor_verified': record['mentor_verified'],
                        'facilitator_verified': record['facilitator_verified'],
                        'color': self.ATTENDANCE_COLORS.get(record['type'], '#e5e7eb'),
                    })
                else:
                    daily_attendance.append({
                        'date': work_date,
                        'day': work_date.day,
                        'weekday': work_date.strftime('%a'),
                        'code': '-',
                        'type': 'NO_RECORD',
                        'verified': False,
                        'mentor_verified': False,
                        'facilitator_verified': False,
                        'color': self.ATTENDANCE_COLORS['NO_RECORD'],
                    })
            
            learner_entry = {
                'placement': placement,
                'learner': placement.learner,
                'learner_name': placement.learner.get_full_name(),
                'learner_id': placement.learner.sa_id_number or placement.learner.learner_number,
                'host': placement.host.company_name if placement.host else '-',
                'mentor': f"{placement.mentor.first_name} {placement.mentor.last_name}" if placement.mentor else '-',
                'daily_attendance': daily_attendance,
                'summary': summary,
                'existing_stipend': existing_stipend,
                'verification_flag': self._get_verification_flag(summary),
            }
            learner_data.append(learner_entry)
            
            # Update totals
            totals['learner_count'] += 1
            totals['total_present'] += summary['type_counts']['PRESENT']
            totals['total_absent'] += summary['type_counts']['ABSENT'] + summary['missed_logging']
            totals['total_leave'] += (
                summary['type_counts']['ANNUAL'] +
                summary['type_counts']['SICK'] +
                summary['type_counts']['FAMILY']
            )
            totals['total_payable_days'] += summary['stipend']['payable_days']
            totals['total_stipend'] += summary['stipend']['net_amount']
            
            # Verification categorization
            if summary['verification']['percentage'] >= 100:
                totals['fully_verified'] += 1
            elif summary['verification']['percentage'] > 0:
                totals['partially_verified'] += 1
            else:
                totals['unverified'] += 1
        
        return {
            'project': self.not_project,
            'period': {
                'year': self.year,
                'month': self.month,
                'month_name': calendar.month_name[self.month],
                'start_date': self.month_start,
                'end_date': self.month_end,
            },
            'working_days': self.working_days,
            'total_working_days': self.total_working_days,
            'learners': learner_data,
            'totals': totals,
            'legend': {
                'codes': self.ATTENDANCE_CODES,
                'colors': self.ATTENDANCE_COLORS,
            },
            'generated_at': timezone.now(),
        }
    
    def _get_verification_flag(self, summary: Dict) -> str:
        """
        Get a verification status flag for display.
        Returns: 'verified', 'partial', or 'unverified'
        """
        pct = summary['verification']['percentage']
        if pct >= 100:
            return 'verified'
        elif pct > 0:
            return 'partial'
        return 'unverified'
    
    def generate_pdf(self) -> bytes:
        """
        Generate a PDF version of the attendance register.
        
        Returns:
            PDF content as bytes
        """
        if not WEASYPRINT_AVAILABLE:
            raise ImportError(
                "WeasyPrint is required for PDF generation. "
                "Install it with: pip install weasyprint"
            )
        
        data = self.get_register_data()
        
        # Render HTML template
        html_content = render_to_string(
            'reports/monthly_attendance_register.html',
            {'register': data}
        )
        
        # Generate PDF
        html = HTML(string=html_content)
        pdf_bytes = html.write_pdf()
        
        return pdf_bytes
    
    def generate_excel(self) -> bytes:
        """
        Generate an Excel version of the attendance register.
        
        Returns:
            Excel file content as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install it with: pip install openpyxl"
            )
        
        data = self.get_register_data()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Attendance {data['period']['month_name'][:3]} {data['period']['year']}"
        
        # Styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws['A1'] = f"Monthly Attendance Register - {data['period']['month_name']} {data['period']['year']}"
        ws['A1'].font = title_font
        ws.merge_cells('A1:H1')
        
        ws['A2'] = f"Project: {data['project'].reference_number} - {data['project'].title}"
        ws.merge_cells('A2:H2')
        
        ws['A3'] = f"Generated: {data['generated_at'].strftime('%Y-%m-%d %H:%M')}"
        ws.merge_cells('A3:H3')
        
        # Header row
        row = 5
        headers = ['#', 'Learner Name', 'ID Number', 'Host Employer']
        
        # Add day columns
        for work_date in data['working_days']:
            headers.append(work_date.day)
        
        # Add summary columns
        headers.extend([
            'Present', 'Leave', 'Absent', 'Missing',
            'Verified %', 'Daily Rate', 'Payable Days', 'Net Stipend'
        ])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        row = 6
        for idx, learner in enumerate(data['learners'], 1):
            col = 1
            
            # Learner info
            ws.cell(row=row, column=col, value=idx).border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=learner['learner_name']).border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=learner['learner_id']).border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=learner['host']).border = thin_border
            col += 1
            
            # Daily attendance
            for day_data in learner['daily_attendance']:
                cell = ws.cell(row=row, column=col, value=day_data['code'])
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # Color coding
                if day_data['type'] == 'PRESENT':
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif day_data['type'] == 'ABSENT':
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif day_data['type'] == 'NO_RECORD':
                    cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                elif day_data['type'] in ['ANNUAL', 'SICK', 'FAMILY']:
                    cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
                
                # Add verification indicator
                if day_data['verified']:
                    cell.font = Font(bold=True)
                
                col += 1
            
            # Summary columns
            summary = learner['summary']
            ws.cell(row=row, column=col, value=summary['type_counts']['PRESENT']).border = thin_border
            col += 1
            leave_total = (
                summary['type_counts']['ANNUAL'] +
                summary['type_counts']['SICK'] +
                summary['type_counts']['FAMILY']
            )
            ws.cell(row=row, column=col, value=leave_total).border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=summary['type_counts']['ABSENT']).border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=summary['missed_logging']).border = thin_border
            col += 1
            
            # Verification percentage with flag
            cell = ws.cell(
                row=row, column=col, 
                value=f"{summary['verification']['percentage']}%"
            )
            cell.border = thin_border
            if summary['verification']['percentage'] < 100:
                cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            col += 1
            
            # Stipend info
            ws.cell(row=row, column=col, value=float(summary['stipend']['daily_rate'])).border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=summary['stipend']['payable_days']).border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=float(summary['stipend']['net_amount'])).border = thin_border
            col += 1
            
            row += 1
        
        # Totals row
        row += 1
        ws.cell(row=row, column=1, value='TOTALS').font = header_font
        
        totals_start_col = 4 + len(data['working_days']) + 1
        ws.cell(row=row, column=totals_start_col, value=data['totals']['total_present']).font = header_font
        ws.cell(row=row, column=totals_start_col + 1, value=data['totals']['total_leave']).font = header_font
        ws.cell(row=row, column=totals_start_col + 2, value=data['totals']['total_absent']).font = header_font
        ws.cell(row=row, column=totals_start_col + 6, value=data['totals']['total_payable_days']).font = header_font
        ws.cell(row=row, column=totals_start_col + 7, value=float(data['totals']['total_stipend'])).font = header_font
        
        # Legend sheet
        legend_ws = wb.create_sheet(title='Legend')
        legend_ws['A1'] = 'Attendance Codes'
        legend_ws['A1'].font = title_font
        
        legend_row = 3
        for code, label in [
            ('P', 'Present'),
            ('AL', 'Annual Leave'),
            ('SL', 'Sick Leave'),
            ('FL', 'Family Responsibility Leave'),
            ('UL', 'Unpaid Leave'),
            ('PH', 'Public Holiday'),
            ('A', 'Absent Without Leave'),
            ('S', 'Suspended'),
            ('-', 'No Record (Missed Logging)'),
        ]:
            legend_ws.cell(row=legend_row, column=1, value=code).font = Font(bold=True)
            legend_ws.cell(row=legend_row, column=2, value=label)
            legend_row += 1
        
        legend_row += 2
        legend_ws.cell(row=legend_row, column=1, value='Verification').font = header_font
        legend_row += 1
        legend_ws.cell(row=legend_row, column=1, value='Bold text = Dual verified (Mentor + Facilitator)')
        legend_row += 1
        legend_ws.cell(row=legend_row, column=1, value='Yellow highlight in Verified % = Incomplete verification')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        
        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def create_deliverable_record(self, pdf_content: bytes = None, excel_content: bytes = None):
        """
        Create a NOTDeliverable record with the generated register attached.
        
        Args:
            pdf_content: PDF file content (optional)
            excel_content: Excel file content (optional)
        
        Returns:
            Created NOTDeliverable instance
        """
        from core.models import NOTDeliverable
        
        month_name = calendar.month_name[self.month]
        
        # Check for existing deliverable first
        existing = NOTDeliverable.objects.filter(
            training_notification=self.not_project,
            deliverable_type='REPORT',
            title=f"Attendance Register - {month_name} {self.year}",
        ).first()
        
        if existing:
            deliverable = existing
            deliverable.status = 'COMPLETED'
            deliverable.completed_date = date.today()
            deliverable.save()
        else:
            deliverable = NOTDeliverable.objects.create(
                training_notification=self.not_project,
                deliverable_type='REPORT',
                title=f"Attendance Register - {month_name} {self.year}",
                description=f"Monthly attendance register for {month_name} {self.year}",
                due_date=self.month_end,
                status='COMPLETED',
                completed_date=date.today(),
            )
        
        # Attach files if provided
        if pdf_content:
            filename = f"attendance_register_{self.not_project.reference_number}_{self.year}_{self.month:02d}.pdf"
            deliverable.attachments.save(filename, ContentFile(pdf_content))
        
        return deliverable
    
    @classmethod
    def setup_recurring_deliverables(cls, training_notification):
        """
        Set up recurring monthly attendance register deliverables for a NOT project.
        Creates a parent deliverable with monthly recurrence.
        
        Args:
            training_notification: TrainingNotification instance
        """
        from core.models import NOTDeliverable
        
        # Calculate first due date (end of first full month after project start)
        start_date = training_notification.planned_start_date or date.today()
        if start_date.day > 1:
            # Move to first of next month
            if start_date.month == 12:
                first_due = date(start_date.year + 1, 1, 1)
            else:
                first_due = date(start_date.year, start_date.month + 1, 1)
        else:
            first_due = start_date
        
        # Get last day of that month
        _, last_day = calendar.monthrange(first_due.year, first_due.month)
        first_due = date(first_due.year, first_due.month, last_day)
        
        # End date for recurrence (project end date or 12 months)
        end_date = training_notification.planned_end_date
        if not end_date:
            end_date = date(first_due.year + 1, first_due.month, last_day)
        
        # Create or update the parent recurring deliverable
        parent, created = NOTDeliverable.objects.update_or_create(
            training_notification=training_notification,
            deliverable_type='REPORT',
            title='Monthly Attendance Register',
            is_recurring=True,
            parent_deliverable__isnull=True,  # Only match parent templates
            defaults={
                'description': 'Monthly attendance register showing learner attendance, verification status, and stipend calculations.',
                'due_date': first_due,
                'recurrence_type': 'MONTHLY',
                'recurrence_end_date': end_date,
                'responsible_department': 'LOGISTICS',
                'submit_to': 'Project Manager / Funder',
                'status': 'PENDING',
            }
        )
        
        # Generate recurring instances if newly created
        if created:
            parent.generate_recurring_instances()
        
        return parent
